import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

#import testapp.procesos_bd as proc_bd
import docode_managebd.procesos_bd as proc_bd

from datetime import datetime
from django.http import HttpResponse

# Metodo para leer un archivo de excel y guardar en el Modelo
def leerExcel(excel_file,modelo):
    respuesta = {
        'resp' : 1,
        'mensaje' : "",
    }

    campos =  proc_bd.obtener_campos(modelo)
    campos.pop(0)
    campos_cont = 1
    dict_cont = 0

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb['datos']

    # iterating over the rows and
    # getting value from each cell in row
    for row in worksheet.iter_rows():
        dict_cont = 0
        objeto = modelo()
        for cell in row:
            dato = cell.value
            try:
                if campos_cont >= len(campos):
                    campo_info = campos[dict_cont]
                    if dato == "" or dato == None:
                        dato == None
                    elif campo_info['tipo'] == 'ForeignKey':
                        querySet = campo_info['qdata']
                        objetoR = querySet.model.objects.filter(id=int(dato)).first()
                        name = campo_info['nombre']
                        setattr(objeto,name,objetoR)
                        a = 0
                    elif campo_info['tipo'] == 'DateField':
                        if type(dato) == datetime:
                            objeto.__dict__[campo_info['nombre']] = dato
                            a = 0
                        elif type(dato)!=datetime:                        
                            objeto.__dict__[campo_info['nombre']] = obtenerFecha(dato)
                    elif campo_info['tipo'] == 'FileField':
                        None
                    elif campo_info['tipo'] != 'DateField':
                        objeto.__dict__[campo_info['nombre']] = dato
                    dict_cont += 1
                campos_cont += 1

            except Exception as e:
                respuesta['resp'] = 0
                respuesta['mensaje'] = "Error: " + str(e)
        # No se registran los datos de encabezado(header)
        if campos_cont > len(campos):
            objeto.save()

    return response

# Metodo para parseart dato string a datetime object
def  obtenerFecha(dato):
        dato = '/'.join(dato.split('-'))
        dato = datetime.strptime(dato, "%d/%m/%Y")

# Metod para obtener el layout de un Modelo en formato de Excel
def obtenerLayout(modelo):
    campos =  proc_bd.obtener_campos(modelo)
    campos.pop(0)
    sheet_index = 0
    
    # Se inicia el objeto response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename='+modelo.__name__+'-{date}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'datos'

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True, color='ffffff')
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='0aa6d4'),
        right=Side(border_style='medium', color='0aa6d4'),
    )

    try:

        # Define the titles for columns
        columns = []
        for campo in campos:
            if campo['foreign'] == True:
                columns.append(campo['nombre'] + "(id)")
            # Omitir los campos de tipo FileField
            elif campo['tipo'] == "FileField":
                None
            else:
                columns.append(campo['nombre'])
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = PatternFill(start_color='08c7ff', end_color='08c7ff', fill_type = 'solid')
        
        for campo in campos:
            if campo['tipo'] == 'choices':
                allChoices = len(campo['qdata'])
                datos = campo['qdata']
                sheet_index += 1
                workbook.create_sheet(campo['nombre'].upper())
                workbook.active = sheet_index
                worksheet = workbook.active
                choices = campo['qdata']
                columns = []
                for choiceID in range(allChoices):
                    if choiceID == 0:
                        columns.append('id')
                    else:
                        columns.append('opciones')
                    row_num = 1

                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.border = border_bottom
                    cell.alignment = centered_alignment
                    cell.fill = PatternFill(start_color='08c7ff', end_color='08c7ff', fill_type = 'solid')

                for dato in datos:    
                    row_num+=1  
                    i = 0
                    for col_num, column_title in enumerate(columns, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = dato[i]        
                        cell.alignment = centered_alignment  
                        i+=1        
                
            elif campo['foreign'] == True:
                sheet_index += 1
                workbook.create_sheet(campo['nombre'] + '(id)')
                workbook.active = sheet_index
                worksheet = workbook.active
                querySet = campo['qdata']
                camposO =  proc_bd.obtener_campos(querySet.model)
                columns = []
                for campo in camposO:
                    columns.append(campo['nombre'])
                    row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.border = border_bottom
                    cell.alignment = centered_alignment
                    cell.fill = PatternFill(start_color='08c7ff', end_color='08c7ff', fill_type = 'solid')


                row_num=1
                for datos in querySet:
                    m = 0
                    row_num+=1
                    for col_num, column_title in enumerate(columns, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = datos.__dict__[column_title]           
                        cell.alignment = centered_alignment                    
                        m += 1
        workbook.save(response)

    except Exception as e:
        None

    return response
